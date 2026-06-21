import logging
import os
import re
import time

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfparser import PDFParser
from pdfminer.pdftypes import resolve1
from openpyxl.styles import PatternFill

# Many PDFs trigger pdfminer FontBBox warnings; text extraction still succeeds.
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# Saving while Excel (or Preview on macOS) has the file open often raises *Permission denied*.
_SAVE_RETRIES = 10
_SAVE_DELAY_SEC = 0.45

_EXCEL_LOCK_USER_MESSAGE = (
    "Could not save the Excel file because it is open in another program "
    "(usually Microsoft Excel or Preview). Close that workbook window, run "
    "this tool again, then reopen the file to see new rows. Excel does not "
    "reload changes from disk while the file stays open."
)


def _is_file_lock_error(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError):
        return True
    if isinstance(exc, OSError):
        errno = getattr(exc, "errno", None)
        if errno in (1, 13):  # EPERM / EACCES — common when the .xlsx is locked
            return True
        msg = str(exc).lower()
        if "permission" in msg or "denied" in msg:
            return True
    return False


def _save_workbook_with_retries(wb, excel_path: str) -> None:
    last: OSError | None = None
    for _ in range(_SAVE_RETRIES):
        try:
            wb.save(excel_path)
            return
        except OSError as e:
            if not _is_file_lock_error(e):
                raise
            last = e
            time.sleep(_SAVE_DELAY_SEC)
    raise PermissionError(_EXCEL_LOCK_USER_MESSAGE) from last


def _run_with_file_lock_retries(operation) -> None:
    """Run *operation* (no-arg callable) retrying on typical Windows/macOS lock errors."""
    last: OSError | None = None
    for _ in range(_SAVE_RETRIES):
        try:
            operation()
            return
        except OSError as e:
            if not _is_file_lock_error(e):
                raise
            last = e
            time.sleep(_SAVE_DELAY_SEC)
    raise PermissionError(_EXCEL_LOCK_USER_MESSAGE) from last

# Column order for new workbooks and appended rows
EXCEL_COLUMNS = [
    "File",
    "Date",
    "Model",
    "Description",
    "Submitter",
    "Report No.",
    "Label",
    "Batch Size",
]


def _line_is_equipment_description_header(line: str) -> bool:
    """Match label whether the PDF uses ASCII or typographic apostrophe in *l'équipement*."""
    if "Equipment Description" not in line or "Description de l" not in line:
        return False
    return "équipement" in line and ":" in line


def _extract_batch_size(full_text: str) -> str | None:
    """Batch size from *Batch Size / Nombre d'échantillon*, or fallbacks when OCR omits the digit."""
    # Digit on the **same line** as the colon (do not let \\s* cross newlines — would grab e.g. *4.* from *4. Comp*)
    m = re.search(
        r"Batch Size\s*/\s*Nombre d['\u2019]?\s*echantillons?\s*:\s*(\d+)\s*(?:\n|$)",
        full_text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1)
    m = re.search(
        r"Batch Size\s*\|\s*Nombre d['\u2019]?\s*échantillons?\s*:\s*(\d+)\s*(?:\n|$)",
        full_text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1)
    # Value on the line after *Nombre d'échantillon :* only if that line is digits-only
    m = re.search(
        r"Nombre d['\u2019]?\s*echantillons?\s*:\s*\n\s*(\d+)\s*(?:\n|$)",
        full_text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1)
    # Many layouts omit the batch digit in the text stream; *Sample Size / …* often matches the same box value
    m = re.search(r"Sample Size\s*/\s*(\d+)", full_text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"Sample Size\s*\|\s*Taille de l['\u2019]?\s*échantillon\s*:\s*(\d+)", full_text, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


_LABEL_FROM_TEXT_RE = re.compile(r"^C-\s*\d+$", re.IGNORECASE)


def _decode_pdf_field_value(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, dict):
        return None
    if isinstance(v, bytes):
        s = v.decode("latin-1", "replace").strip()
    else:
        s = str(v).strip()
    if not s:
        return None
    if s.startswith("/'") and s.endswith("'"):
        return None
    if s.startswith("{'") and "Sig" in s:
        return None
    return s


def _acro_norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _iter_leaf_acroform_widgets(pdf_path: str):
    with open(pdf_path, "rb") as fp:
        parser = PDFParser(fp)
        doc = PDFDocument(parser)
        catalog = resolve1(doc.catalog)
        if not catalog:
            return
        acro = resolve1(catalog.get("AcroForm"))
        if not acro:
            return
        fields = resolve1(acro.get("Fields"))
        if not fields:
            return

        def walk(field_ref):
            f = resolve1(field_ref)
            if not f:
                return
            kids = f.get("Kids")
            if kids:
                for k in resolve1(kids):
                    yield from walk(k)
                return
            yield f

        for ref in fields:
            yield from walk(ref)


def _extract_from_acroform(pdf_path: str) -> dict[str, str | None]:
    """Read filled AcroForm values (new Word→PDF forms often omit them from *extract_text*)."""
    empty = {
        "Date": None,
        "Model": None,
        "Description": None,
        "Submitter": None,
        "Report No.": None,
        "Label": None,
        "Batch Size": None,
    }
    try:
        widgets = list(_iter_leaf_acroform_widgets(pdf_path))
    except Exception:
        return empty
    if not widgets:
        return empty
    out = dict(empty)
    label_candidates: list[str] = []
    for w in widgets:
        raw_name = resolve1(w.get("T")) if w.get("T") else ""
        if isinstance(raw_name, bytes):
            raw_name = raw_name.decode("latin-1", "replace")
        val = w.get("V")
        val = resolve1(val) if val else None
        sval = _decode_pdf_field_value(val)
        if not sval:
            continue
        nn = _acro_norm_name(raw_name)
        if _LABEL_FROM_TEXT_RE.match(sval):
            label_candidates.append(sval)
        if ("entreprise" in nn or re.match(r"^company\b", nn)) and "signature" not in nn:
            out["Submitter"] = sval
        elif "equipment description" in nn and "row" not in nn:
            out["Description"] = sval
        elif "model" in nn and ("modèle" in nn or "modele" in nn) and "row" not in nn:
            out["Model"] = sval
        elif "date2_af_date" in nn or re.search(r"\bmdy\b|mm/jj/aaaa", nn):
            if re.search(r"\d", sval):
                out["Date"] = sval
        elif re.search(r"report\s*#|no\s+du\s+rapport|no\.\s*rapport", nn):
            out["Report No."] = sval
        elif nn == "batch size" or (nn.startswith("batch size") and "sample" not in nn):
            if re.fullmatch(r"\d+", sval.strip()):
                out["Batch Size"] = sval.strip()
    if label_candidates:
        out["Label"] = label_candidates[0]
    return out


def _report_and_company_from_filename(pdf_path: str) -> tuple[str | None, str | None]:
    """*SI#2026-WZ-1084 smart flex Lighting NEW FORMS copy.pdf* → report + company tail."""
    stem = re.sub(r"\.pdf\s*$", "", os.path.basename(pdf_path), flags=re.IGNORECASE)
    m = re.match(r"(?i)(?:SI#)?(\d{4}-\w+-\d+)\s+(.+)$", stem.strip())
    if not m:
        return None, None
    report, rest = m.group(1), m.group(2).strip()
    rest = re.sub(r"(?i)(\s+new forms.*|\s+copy.*)$", "", rest).strip()
    company = rest or None
    return report, company


def _description_from_equipment_line(line: str) -> str | None:
    m = re.search(
        r"Equipment Description\s*(?:/\s*|\|\s*)\s*Description de l['\u2019]?équipement\s*:\s*(.+)$",
        line,
        re.IGNORECASE,
    )
    if not m:
        return None
    frag = m.group(1).strip()
    # Stop if another bilingual label starts mid-line (rare merged export)
    cut = re.split(
        r"(?i)\s+(?:Address\s*(?:/\s*|\|\s*)\s*Adresse|Client\s*#|Contact\s*(?:/\s*|\|\s*))",
        frag,
        maxsplit=1,
    )
    return cut[0].strip() or None


def _description_from_following_address_line(next_line: str) -> str | None:
    for addr in (
        "Address / Adresse:",
        "Address / Adresse :",
        "Address | Adresse:",
        "Address | Adresse :",
    ):
        idx = next_line.find(addr)
        if idx == -1:
            continue
        description = next_line[idx + len(addr) :].strip()
        description = re.sub(r"^\d+\s+\S+[^,]*,", "", description).strip()
        return description or None
    return None


def _extract_submitter_from_text(text: str) -> str | None:
    m = re.search(
        r"Submitter\s+/\s+Réquérant:\s+([^\n]+?)\s+Page",
        text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1).strip()
    m = re.search(
        r"(?m)^[^\n]*Company\s*\|\s*Entreprise\s*:[ \t]*([^\n]*?)\s+Page\s*:",
        text,
        re.IGNORECASE,
    )
    if m:
        s = m.group(1).strip()
        return s or None
    return None


def _extract_date_from_text(text: str) -> str | None:
    m = re.search(r"M.D.Y\s+/\s+M.J.A:\s+([0-9/]+)", text)
    if m:
        return m.group(1)
    m = re.search(
        r"M.D.Y\s*\|\s*mm/jj/aaaa\s*:[ \t]*([0-9/]+)",
        text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1)
    return None


def _is_plausible_intertek_report_no(s: str | None) -> bool:
    if not s:
        return False
    return bool(re.fullmatch(r"(?i)\d{4}-\w+-\d+", s.strip()))


def _extract_model_from_text(text: str) -> str | None:
    m = re.search(r"Model\s+/\s+Modèle:\s+(\S+)", text)
    if m:
        return m.group(1)
    # Pipe layout: value must be on the same line; do not use \s* after the colon (it would span to *Tel* on the next line).
    m = re.search(
        r"(?m)^[^\n]*Model\s*\|\s*Modèle\s*:[ \t]*([^\s\n]+)",
        text,
        re.IGNORECASE,
    )
    if m:
        return m.group(1)
    return None


def _extract_report_from_text(text: str) -> str | None:
    m = re.search(r"Report No[./]*\s+No\.\s+Rapport:\s*(\S+)", text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(
        r"(?m)^[^\n]*Report\s*#\s*\|\s*No\s+du\s+rapport\s*:[ \t]*([^\s\n]*)",
        text,
        re.IGNORECASE,
    )
    if m:
        v = (m.group(1) or "").strip()
        return v or None
    return None


def _pick_report_no(*candidates: str | None) -> str | None:
    for c in candidates:
        if _is_plausible_intertek_report_no(c):
            return str(c).strip()
    return None


def _header_row_looks_like_serial_label_table(row) -> bool:
    if not row:
        return False
    cells = [((c or "").strip()).upper() for c in row if c is not None]
    joined = " ".join(cells)
    return "SERIAL" in joined and ("LABEL" in joined or "ETIQUETTE" in joined)


def _extract_label_from_pdf(pdf) -> str | None:
    """First LABEL / ETIQUETTE cell on the row numbered *1* in the serial-number table(s)."""
    for page in pdf.pages:
        for table in page.extract_tables() or []:
            header_idx = None
            for i, row in enumerate(table):
                if row and _header_row_looks_like_serial_label_table(row):
                    header_idx = i
                    break
            if header_idx is None:
                continue
            header = table[header_idx]
            label_col = None
            for j, cell in enumerate(header):
                if cell is None:
                    continue
                u = (cell or "").upper()
                if "LABEL" in u or "ETIQUETTE" in u:
                    label_col = j
                    break
            if label_col is None:
                continue
            for row in table[header_idx + 1 :]:
                if not row or len(row) <= label_col:
                    continue
                if (row[0] or "").strip() != "1":
                    continue
                lab = row[label_col]
                if lab is None:
                    continue
                s = str(lab).strip()
                if s:
                    return s
    return None


def process_pdf(pdf_path):
    """Extract labeled fields from the PDF (header fields from page 1; batch/label from whole file)."""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        first_page = pdf.pages[0]
        text = first_page.extract_text()
        if text is None:
            raise ValueError("No text found on the first page.")

        lines = text.split("\n")

        description = None
        for i, line in enumerate(lines):
            if _line_is_equipment_description_header(line):
                description = _description_from_equipment_line(line)
                if description is None and i + 1 < len(lines):
                    description = _description_from_following_address_line(lines[i + 1])
                break

        date = _extract_date_from_text(text)
        model = _extract_model_from_text(text)
        submitter = _extract_submitter_from_text(text)
        report_text = _extract_report_from_text(text)

        batch_size = _extract_batch_size(full_text)
        label = _extract_label_from_pdf(pdf)

        form = _extract_from_acroform(pdf_path)
        report_from_name, company_from_name = _report_and_company_from_filename(pdf_path)

        def coalesce(primary, key: str) -> str | None:
            if primary:
                return primary
            v = form.get(key)
            return v if v else None

        date = coalesce(date, "Date")
        model = coalesce(model, "Model")
        submitter = coalesce(submitter, "Submitter")
        description = coalesce(description, "Description")
        batch_size = batch_size or form.get("Batch Size")
        label = label or form.get("Label")

        report_num = _pick_report_no(
            report_text, form.get("Report No."), report_from_name
        )
        submitter = submitter or company_from_name

        return {
            "File": pdf_path,
            "Date": date,
            "Model": model,
            "Description": description,
            "Submitter": submitter,
            "Report No.": report_num,
            "Label": label,
            "Batch Size": batch_size,
        }


def _cell_nonempty(cell) -> bool:
    v = cell.value
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


def _argb_to_8(s: str) -> str:
    """Normalize openpyxl *rgb* string to 8-char ARGB uppercase."""
    r = str(s).strip().upper()
    if len(r) == 6:
        return "FF" + r
    return r


def _cell_has_visible_fill(cell) -> bool:
    """True if the cell has a non-default background (pattern or non-white solid)."""
    f = cell.fill
    pt = f.patternType
    if pt in (None, "none"):
        return False
    if pt == "solid":
        fc = f.fgColor
        if fc.type == "rgb" and fc.rgb:
            a8 = _argb_to_8(fc.rgb)
            # Transparent or opaque white — treat as no highlight
            if a8 == "00000000" or a8[2:] == "FFFFFF":
                return False
            return True
        if fc.type == "indexed" and fc.indexed is not None:
            # 64 / 65 = automatic fg/bg per ECMA-376
            return fc.indexed not in (64, 65)
        if fc.type == "theme":
            return True
        return False
    # gray125 banding, gradients, etc.
    return True


def _is_legacy_app_workbook(ws) -> bool:
    """Workbooks created by this app use a header row starting with *File*."""
    if ws.max_row < 1:
        return False
    return ws.cell(row=1, column=1).value == "File"


def _normalized_sheet_name(sheet_name: str | None) -> str | None:
    if sheet_name is None:
        return None
    s = str(sheet_name).strip()
    return s or None


def _resolve_worksheet(wb, sheet_name: str | None):
    """Return a worksheet by name, or the workbook's active sheet if *sheet_name* is empty."""
    name = _normalized_sheet_name(sheet_name)
    if name is not None:
        if name not in wb.sheetnames:
            available = ", ".join(repr(s) for s in wb.sheetnames)
            raise ValueError(f"No sheet named {name!r}. Available: {available}")
        return wb[name]
    return wb.active


def _first_row_b_c_d_empty_no_fill(ws) -> int:
    """First row where B, C, and D are empty and none has a background fill."""
    max_r = max(ws.max_row, 1)
    for r in range(1, max_r + 1):
        b_cell = ws.cell(row=r, column=2)
        c_cell = ws.cell(row=r, column=3)
        d_cell = ws.cell(row=r, column=4)
        b_ok = not _cell_nonempty(b_cell) and not _cell_has_visible_fill(b_cell)
        c_ok = not _cell_nonempty(c_cell) and not _cell_has_visible_fill(c_cell)
        d_ok = not _cell_nonempty(d_cell) and not _cell_has_visible_fill(d_cell)
        if b_ok and c_ok and d_ok:
            return r
    return max_r + 1


def _serial_number_from_label(label: str | None) -> int | None:
    """Strip leading *C-* and parse the numeric part (e.g. *C-2307970* → 2307970)."""
    if label is None:
        return None
    s = str(label).strip()
    s = re.sub(r"^[Cc]-\s*", "", s)
    if not s.isdigit():
        return None
    return int(s)


def _batch_size_positive(batch: str | int | None) -> int:
    if batch is None:
        return 1
    try:
        return max(1, int(str(batch).strip()))
    except (ValueError, TypeError):
        return 1


def _label_serial_columns_e_f(
    label: str | None, batch_size: str | int | None
) -> tuple[int | None, int | None]:
    """E = start serial (after *C-*), F = E + batch_size - 1 (same as E when batch is 1)."""
    base = _serial_number_from_label(label)
    if base is None:
        return None, None
    n = _batch_size_positive(batch_size)
    return base, base + n - 1


def _append_template_row(ws, new_data: dict, column_b_date: str | None) -> None:
    """Master sheet: **A** = Report No. if A is empty; B if *column_b_date*; C/D; E/F; J=batch; K=file."""
    row = _first_row_b_c_d_empty_no_fill(ws)
    blank_fill = PatternFill()
    a_cell = ws.cell(row=row, column=1)
    if not _cell_nonempty(a_cell):
        rep = new_data.get("Report No.")
        if rep is not None and str(rep).strip():
            a_cell.value = str(rep).strip()
            a_cell.fill = blank_fill
    if column_b_date is not None and str(column_b_date).strip():
        ws.cell(row=row, column=2, value=str(column_b_date).strip())
    c_cell = ws.cell(row=row, column=3)
    d_cell = ws.cell(row=row, column=4)
    c_cell.value = new_data.get("Submitter")
    d_cell.value = new_data.get("Model")
    # Ensure new values do not keep table banding / highlight fills
    c_cell.fill = blank_fill
    d_cell.fill = blank_fill
    e_val, f_val = _label_serial_columns_e_f(
        new_data.get("Label"), new_data.get("Batch Size")
    )
    ws.cell(row=row, column=5, value=e_val)
    ws.cell(row=row, column=6, value=f_val)
    ws.cell(row=row, column=10, value=new_data.get("Batch Size"))
    file_path = new_data.get("File")
    if file_path:
        ws.cell(row=row, column=11, value=os.path.basename(str(file_path)))


def create_master_list_workbook(excel_path: str, sheet_title: str = "Sheet1") -> None:
    """Create an empty .xlsx with the **master list** header row (same columns *Process* writes).

    Row 1: **A** Report No., **B** Date, **C** Submitter, **D** Model, **E**/**F** serial range,
    **J** Batch Size, **K** File (PDF base name). Cell **A1** is not ``File``, so *Process*
    treats the workbook as a master list, not the legacy export layout.
    """
    wb = Workbook()
    ws = wb.active
    title = (sheet_title or "Sheet1").strip() or "Sheet1"
    ws.title = title[:31]
    headers = {
        1: "Report No.",
        2: "Date",
        3: "Submitter",
        4: "Model",
        5: "Serial start",
        6: "Serial end",
        10: "Batch Size",
        11: "File",
    }
    for col, label in headers.items():
        ws.cell(row=1, column=col, value=label)
    out_dir = os.path.dirname(os.path.abspath(excel_path)) or "."
    os.makedirs(out_dir, exist_ok=True)
    wb.save(excel_path)
    wb.close()


def update_excel(
    new_data,
    excel_path,
    column_b_date: str | None = None,
    sheet_name: str | None = None,
):
    """Append one row to an Excel file (create if missing).

    If *excel_path* does not exist, a **master-list** workbook is created (same
    layout as **Create New Excel File** in the GUI), then the first data row is written.

    If the workbook is a legacy export from this tool (header *File* in A1),
    rows are appended using the original column layout.

    Otherwise the file is treated as an existing master list: the **first row
    where B, C, and D are all empty and none has a background fill** receives
    **Report No.** in **A** when **A** is empty (otherwise **A** is left unchanged),
    **Submitter** in **C**, **Model** in **D**, serial range **E** (start, after
    stripping a leading *C-*) and **F** (end = start + batch size − 1), **Batch
    Size** in **J**, and PDF **file name** in **K** (any fill on A/C/D for those
    writes is cleared). If no such row exists through the used range, a new row is added
    after the last row.
    Column **B** is written only when *column_b_date* is a non-empty string;
    otherwise cell B on that row is left blank.

    *sheet_name*: worksheet to use. If omitted or blank, the workbook's **active**
    sheet is used (the tab that was selected when the file was saved in Excel).

    If the workbook is open in Excel (or Preview on macOS), saving may be blocked;
    the implementation retries briefly, then raises with a clear message if the
    file is still locked.
    """
    sheet_name = _normalized_sheet_name(sheet_name)

    if not os.path.exists(excel_path):
        create_master_list_workbook(excel_path)
        wb_new = load_workbook(excel_path)
        try:
            ws_new = _resolve_worksheet(wb_new, sheet_name)
            _append_template_row(ws_new, new_data, column_b_date)
            _save_workbook_with_retries(wb_new, excel_path)
        finally:
            wb_new.close()
        return

    wb = None
    try:
        wb = load_workbook(excel_path)
        ws = _resolve_worksheet(wb, sheet_name)
        target_sheet = ws.title

        if _is_legacy_app_workbook(ws):
            wb.close()
            wb = None
            df = pd.read_excel(excel_path, sheet_name=target_sheet, engine="openpyxl")
            new_data_df = pd.DataFrame([new_data])
            df = pd.concat([df, new_data_df], ignore_index=True)

            def _legacy_save():
                with pd.ExcelWriter(
                    excel_path,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                ) as writer:
                    df.to_excel(writer, sheet_name=target_sheet, index=False)

            _run_with_file_lock_retries(_legacy_save)
            return

        _append_template_row(ws, new_data, column_b_date)
        _save_workbook_with_retries(wb, excel_path)
    finally:
        if wb is not None:
            wb.close()


if __name__ == "__main__":
    import argparse
    from pathlib import Path

    parser = argparse.ArgumentParser(description="Append one row per fileTest*.pdf to output.xlsx.")
    parser.add_argument(
        "--column-b-date",
        metavar="TEXT",
        default=None,
        help="If set, written to column B on the master sheet; if omitted, column B is left blank.",
    )
    parser.add_argument(
        "--sheet",
        metavar="NAME",
        default=None,
        help="Worksheet name to append to (default: active sheet, or first sheet for legacy layout).",
    )
    args = parser.parse_args()
    column_b_date = args.column_b_date
    sheet_name = args.sheet

    root = Path(__file__).resolve().parent
    excel_path = root / "output.xlsx"
    pdf_paths = sorted(root.glob("fileTest*.pdf"))
    if not pdf_paths:
        raise SystemExit(f"No fileTest*.pdf files found in {root}")

    for pdf_path in pdf_paths:
        new_data = process_pdf(str(pdf_path))
        update_excel(
            new_data,
            str(excel_path),
            column_b_date=column_b_date,
            sheet_name=sheet_name,
        )

    print(f"Data extraction and update completed → {excel_path}")
